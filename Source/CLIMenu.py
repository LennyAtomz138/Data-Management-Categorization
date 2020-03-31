"""
CLI Menu is used to display the DMCT menu at application startup.
It is called by Main.py.
"""
from Source import TextractPDFandDOCXVersion, FileHandle, TextractPNGJPGVersion
from Source.ManageBuckets import ViewAndSelectBucketFiles
from openpyxl import Workbook, load_workbook


def MainMenu():
    """
    The Main Menu for the DMCT program that contains submenus for user navigation.
    :return:
    """

    print("*=*" * 14)
    print("Database Management Categorization Tool")
    # print("\t" * 2, "(DMCT)") <- This version shifts the logo over to fit CLI a little better.
    print("\t" * 3, " " * 2, "(DMCT)")
    print("*=*" * 14)

    # TODO: Add a constant notifier that shows what bucket user is in.
    # TODO: Add intermediate messages in-between steps to keep user informed.
    # TODO: Add nested feature to Main Menu s.t. user can view bucket, then select bucket, etc.
    while True:
        print("=" * 14, "Main Menu", "=" * 16)
        print("",
              "1 - Input Keyword(s) and Parse Documents\n",
              "2 - View & Select Bucket\n",
              "3 - View & Select Bucket Files\n",
              "4 - Test Excel Tagging\n",
              "5 - Test Load Excel\n",
              "0 - Exit DMCT")
        print("=" * 41)
        user_input = int(input("Enter Number: "))

        try:
            if user_input < 0 or user_input > 5:
                raise ValueError
            elif user_input == 0:
                print("Exiting the Data Management Categorization Tool")
                break
            elif user_input == 1:
                print("\n")
                GetUserKeywords()
            elif user_input == 2:
                print("\n")
                # TODO: Add ViewAndSelectBucket() here...
            elif user_input == 3:
                print("\n")
                ViewAndSelectBucketFiles()
            elif user_input == 4:
                print("\n")
                TestExcelTagging()
            elif user_input == 5:
                print("\n")
                TestExcelLoading()
            else:
                print("Invalid input: Please try again.")
        except ValueError:
            print("Invalid integer. Please enter valid input.")


def GetUserKeywords():
    """
    Prompts user for keywords and stores them in an array.
    Displays the array to the screen upon completion.
    :return:
    """
    keyword_list = []
    keyword_counter = 0

    print("=" * 8, "Keyword Entry Screen", "=" * 11)
    print("(Input '0' when finished)")
    print("=" * 41)

    while True:
        keyword_counter += 1
        user_input = input("Enter keyword # {counter}: ".format(counter=keyword_counter))
        if user_input.lower() == '0':
            break
        else:
            keyword_list.append(user_input.lower())

    keyword_list.sort()
    print("=" * 41)
    print("You've entered the following keyword(s):\n", keyword_list)
    user_input = int(input("Proceed with document tagging?\n"
                           "Enter 1 for 'Yes' or 0 for 'No': "))
    try:
        if user_input < 0 or user_input > 1:
            raise ValueError
        elif user_input == 0:
            print("Would you like to try again?\n",
                  "1 - Try Again\n",
                  "0 - Quit to Main Menu")
            user_input = int(input("Enter Number: "))
            try:
                if user_input < 0 or user_input > 1:
                    raise ValueError
                elif user_input == 0:
                    print("\n")
                    MainMenu()
                elif user_input == 1:
                    print("=" * 41)
                    print("\n")
                    GetUserKeywords()
                else:
                    print("Invalid input: Please try again.")
            except ValueError:
                print("Invalid integer. Please enter a value between 0 and 1.")
        elif user_input == 1:
            TextractPDFandDOCXVersion.Main(keyword_list)
            #TextractPNGJPGVersion.Main(keyword_list)
        else:
            print("Invalid input: Please try again.")
    except ValueError:
        print("Invalid integer. Please enter a value between 0 and 1.")


def TestExcelTagging():
    """
    Proof of concept function for outputting tagged files into an Excel Document
    3 Test Documents (with pre-defined tags) used for testing
    Output document saved to the Source directory at the moment
    :return:
    """
    print("=" * 8, "Excel Tagging Test", "=" * 8)

    # Test files

    file1 = FileHandle.FileHandle("Planes", "docx")
    file2 = FileHandle.FileHandle("Trains", "pdf")
    file3 = FileHandle.FileHandle("Automobiles", "jpeg")

    file1.addtag("aviation")
    file1.addtag("documentation")

    file2.addtag("locomotive")
    file2.addtag("promotional")

    file3.addtag("automotive")
    file3.addtag("image")

    # List of all files to add to Excel doc

    files = [file1, file2, file3]

    # Set up Excel Doc

    outputbook = Workbook()
    worksheet = outputbook.active
    worksheet['A1'] = "File Name"
    worksheet['B1'] = "File Type"
    worksheet['C1'] = "File Tag 1"
    worksheet['D1'] = "File Tag 2"

    # Add the file names, types, and tags to the sheet

    fileCount = 2
    for file in files:
        worksheet.cell(row=fileCount, column=1, value=file.fileName)
        worksheet.cell(row=fileCount, column=2, value=file.fileType)
        tagCol = 3
        for tag in file.tags:
            worksheet.cell(row=fileCount, column=tagCol, value=tag)
            tagCol += 1
        fileCount += 1

    # Save the doc
    outputbook.save('DCMT_Results.xlsx')


def TestExcelLoading():
    print("=" * 8, "Excel Loading Test", "=" * 8)

    # Path can also be a path to a excel file location.
    path = 'DCMT_Results.xlsx'

    wb = load_workbook(path)
    ws = wb.active

    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            c = ws.cell(row = i, column = j)

            # A check that only exists because not all the headers are labelled.
            if c.value is None:
                break

            print(c.value.center(15), end=" ")
        print('\n')
