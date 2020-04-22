
from Source import FileHandle, CLIMenu
from openpyxl import Workbook, load_workbook

memoryFileName = 'DCMT_Results.xlsx'
entryCount = 0
workbook = Workbook()

def Main(text_dictionary = {}):
    ExcelInitialization()
    while True:
        print("=" * 8, "Excel Manager", "=" * 8)
        print("1 - Print Excel Contents")
        print("2 - Reset Excel Sheet")
        print("0 - Quit to Main Menu\n")
        print("Current Dictionary Contents: \n", text_dictionary)
        ok_to_exit = int(input("Enter Number: "))
        if ok_to_exit < 0 or ok_to_exit > 2:
            print("Invalid input: Please try again.")
            continue
        elif ok_to_exit == 0:
            print("\n")
            return
        elif ok_to_exit == 1:
            PrintContents()
        elif ok_to_exit == 2:
            ExcelReset()
"""A memory check that determines whether an excel file should be created or loaded
    It is run whenever another program makes first access into the ExcelManager"""
def ExcelInitialization():
    global workbook
    try:
        #tries opening Excel memory file
        workbook = load_workbook(memoryFileName)
        activeWorkbook = workbook.active
        print(memoryFileName, " has been loaded\n")
    except:
        #if it doesn't work, it creates the workbook
        workbook = Workbook()
        activeWorkbook = workbook.active
        activeWorkbook['A1'] = "FILE_NAME"
        activeWorkbook['B1'] = "FILE_TYPE"
        activeWorkbook['C1'] = "BUCKET_LOCATION"
        activeWorkbook['D1'] = "FILE_TAG_1"
        activeWorkbook['E1'] = "FILE_TAG_2"
        activeWorkbook['F1'] = "FILE_TAG_3"
        activeWorkbook['G1'] = "FILE_TAG_4"
        activeWorkbook['H1'] = "FILE_TAG_5"
        workbook.save(memoryFileName)
        print(memoryFileName, " has been created\n")
    finally:
        global entryCount
        entryCount = activeWorkbook.max_row - 1
        print("Entry count is ", entryCount)

"""Resets the contents of the excel file to allow for a clean slate"""
def ExcelReset():
    print("=" * 8, "Excel Reset", "=" * 8)
    while True:
        print("Would you like to reset the contents of the Excel sheet?")
        print("1 - Yes")
        print("0 - No\n")
        ok_to_exit = int(input("Enter Number: "))
        if ok_to_exit < 0 or ok_to_exit > 1:
            print("Invalid input: Please try again.")
            continue
        elif ok_to_exit == 0:
            print("\n")
            return
        elif ok_to_exit == 1:
            while True:
                print("Are you sure?")
                print("1 - Confirm")
                print("0 - Cancel\n")
                ok_to_confirm = int(input("Enter Number: "))
                if ok_to_confirm < 0 or ok_to_confirm > 1:
                    print("Invalid input: Please try again.")
                    continue
                elif ok_to_confirm == 0:
                    print("\n")
                    return
                elif ok_to_confirm == 1:
                    global workbook
                    workbook = Workbook()
                    activeWorkbook = workbook.active
                    activeWorkbook['A1'] = "FILE_NAME"
                    activeWorkbook['B1'] = "FILE_TYPE"
                    activeWorkbook['C1'] = "BUCKET_LOCATION"
                    activeWorkbook['D1'] = "FILE_TAG_1"
                    activeWorkbook['E1'] = "FILE_TAG_2"
                    activeWorkbook['F1'] = "FILE_TAG_3"
                    activeWorkbook['G1'] = "FILE_TAG_4"
                    activeWorkbook['H1'] = "FILE_TAG_5"
                    workbook.save(memoryFileName)
                    return

"""Can only be called from the Textract programs, cannot be called from Main Menu
    Adds an scanned file entry to the excel file"""
def AddEntry(text_dictionary = {}):
    ExcelInitialization()
    print("=" * 8, "Add File Entry", "=" * 8)
    global workbook
    workbook = load_workbook(memoryFileName)
    activeWorkbook = workbook.active
    if not text_dictionary: #if the text_dictionary is empty
        print("There are no tags that were scanned for. Returning to Main Menu")
        return 1

    print("Choose at most five keywords that will be applied to the file")
    selected_keys = KeywordSelection(text_dictionary)
    print("The selected keywords are ", selected_keys)
    file = FileHandle.FileHandle(CLIMenu.selected_filename)
    input_row = entryCount + 2
    activeWorkbook.cell(row = input_row, column = 1, value = file.file_name)
    activeWorkbook.cell(row = input_row, column = 2, value = file.file_type)
    activeWorkbook.cell(row = input_row, column = 3, value = file.bucket)
    for i in range(0, len(selected_keys)):
        activeWorkbook.cell(row = input_row, column = i + 4, value = selected_keys[i])
    workbook.save(memoryFileName)
    print("New entry successfully saved in Excel File")
    return selected_keys

"""Can only be called from the ManageBuckets programs, cannot be called from Main Menu
    Checks for a viable file in the excel memory to modify"""
def ModifyEntry(filename, source_bucket, destination_bucket):
    global workbook
    workbook = load_workbook(memoryFileName)
    activeWorkbook = workbook.active
    for row in activeWorkbook.rows:
        if row[0].value is None:
            break
        if (row[0].value == filename) and (row[2].value == source_bucket):
            row[2].value = destination_bucket
            print(filename, " has been moved from ", row[2].value," aka ", source_bucket,  " to ", destination_bucket)
            break
    workbook.save(memoryFileName)
    return



"""The user selects which keywords they've scanned for that they want saved into the excel memory file
    This is due to the five tag limitation"""
def KeywordSelection(text_dictionary = {}):
    selected_keys = []
    keyword_counter = 1
    print("=" * 8, "Keyword Selection", "=" * 8)
    print("(Input '0' when finished)")
    print("Current Dictionary Contents: \n", text_dictionary)
    print("=" * 41)
    while True:
        if keyword_counter > len(text_dictionary) or keyword_counter > 5:
            print("No more entries can be added")
            break

        user_input = input("Enter keyword # {counter}: ".format(counter=keyword_counter))
        if user_input.lower() == '0':
            break
        else:
            if user_input in text_dictionary.keys():
                selected_keys.append(user_input.lower())
                keyword_counter += 1
            else:
                print("Invalid keyword, input a valid keyword")
                print("Current Dictionary Contents: \n", text_dictionary)
    selected_keys.sort()
    return selected_keys

"""Prints the contents of the excel file for viewingr"""
def PrintContents():
    print("=" * 8, "Excel Printed Contents", "=" * 8)
    activeWorkbook = workbook.active

    for i in range(1, activeWorkbook.max_row + 1):
        for j in range(1, 9):
            c = activeWorkbook.cell(row = i, column = j)
            if c.value is None:
                break
            print(c.value.center(15), end=" ")
        print('\n')
